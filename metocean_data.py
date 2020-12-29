"""
Module for the MetoceanData class
By Guillermo Tornero
Metocean & Energy Assessment Department
18/12/2020
"""

import sys
import os
import time
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

import pandas as pd
import numpy as np

import NSS

class MetoceanData:
    """A class to manage store the user configuration settings and read and store the data inputs."""

    # Inisialise the MetoceanData object using the filepath of the configuration file.
    def __init__(self, filepath):
        # Initialise a config attribute which will be a dictionary containing all of the configuration options for the report.
        self.config = {}
        # Initialise a bins attribute which will be a dictionary of lists containing the centre of the different data type bins
        self.bins = {} 
        # Execute the parse_config file to populate the config attribute.
        self.parse_config(filepath)
        # Read and store the data
        self.parse_data()
        # Create sector and bins from data and populate the bins attribute
        self.sectorise()

    def parse_config(self, filepath):
        """parse_config [Parses the 'Config' sheet and stores all configuration parameters in a dictionary self.config.]

        Args:
            filepath ([string]): [full filepath of the config excel file.]
        """

        workbook = load_workbook(filepath, read_only=True)
        # Check if the 'Config' sheet exists in the config file.
        if "Config" in workbook.sheetnames:
            config_sheet = workbook["Config"]
        else:
            # Exit the application if no 'config' sheet exists.
            sys.exit("No 'Config' worksheet found.")

        print("Parsing configuration...", end="")

        self.config["project"] = config_sheet["D5"].value  # Project name [string]
        # Type of variable bin logic <= x < vs < x <=
        if config_sheet["D7"].value == "left" or "right":
            self.config["bin_type"] = config_sheet["D7"].value
        else:
            # Bin type defaults to left is there is an erroneous input in the config file for any reason
            self.config["bin_type"] = "left"

        # method for treating data within wind speed bin for NSS tables
        if config_sheet["D6"].value == "mean" or "median":
            self.config["method"] = config_sheet["D6"].value
        else:
            # Method defaults to median if there is an erroneous input in the config file for any reason
            self.config["method"] = "median"

        # Parsing Config of wind data
        if config_sheet["F9"].value == "ON":
            self.config["wind_status"] = True
        else:
            self.config["wind_status"] = False

        if self.config["wind_status"]:
            self.config["wind_source"] = config_sheet["D10"].value
            self.config["wind_projection"] = config_sheet["D11"].value
            self.config["wind_easting"] = config_sheet["D12"].value
            self.config["wind_northing"] = config_sheet["D13"].value
            self.config["hub_weibull_a"] = config_sheet["D14"].value
            self.config["hub_weibull_k"] = config_sheet["D15"].value
            self.config["hub_height"] = config_sheet["D16"].value
            self.config["10m"] = config_sheet["D17"].value
            self.config["wind_bin_size"] = config_sheet["D18"].value
            self.config["wind_sectors"] = config_sheet["D19"].value

        # Parsing config of wave data
        if config_sheet["F21"].value == "ON":
            self.config["wave_status"] = True
        else:
            self.config["wave_status"] = False

        if self.config["wave_status"]:
            self.config["wave_source"] = config_sheet["D22"].value
            self.config["wave_projection"] = config_sheet["D23"].value
            self.config["wave_easting"] = config_sheet["D24"].value
            self.config["wave_northing"] = config_sheet["D25"].value
            self.config["wave_spectral"] = config_sheet["D26"].value
            self.config["peak_enhancement"] = config_sheet["D27"].value
            self.config["derive_peak_enhancement"] = config_sheet["D28"].value
            self.config["wave_height_bin_size"] = config_sheet["D29"].value
            self.config["wave_period_bin_size"] = config_sheet["D30"].value
            self.config["wave_sectors"] = config_sheet["D31"].value

        # Parsing config of current data
        if config_sheet["F32"].value == "ON":
            self.config["current_status"] = True
        else:
            self.config["current_status"] = False

        if self.config["current_status"]:
            self.config["current_source"] = config_sheet["D34"].value
            self.config["current_projection"] = config_sheet["D35"].value
            self.config["current_easting"] = config_sheet["D36"].value
            self.config["current_northing"] = config_sheet["D37"].value
            self.config["current_bin_size"] = config_sheet["D38"].value
            self.config["current_sectors"] = config_sheet["D39"].value
            self.config["current_components"] = config_sheet["D40"].value

        # Parsing config of seawater data
        if config_sheet["F41"].value == "ON":
            self.config["water_status"] = True
        else:
            self.config["water_status"] = False
        if self.config["water_status"]:
            self.config["water_source"] = config_sheet["D43"].value
            self.config["water_projection"] = config_sheet["D44"].value
            self.config["water_easting"] = config_sheet["D45"].value
            self.config["water_northing"] = config_sheet["D46"].value

        print("Parsing configuration complete!")

    def parse_data(self):
        """parse_data Function to parse the input data .txt files selected by the user.
        Merges all of the input files into a single pandas Dataframe and stores it in a self.data attribute to the MetoceanData class.
        """
        print("Parsing data...", end="")
        # Create a list to store all of the loaded dataframes from the .txt files
        df_list = []
        if self.config["wind_status"]:
            wind_df = self.parse_wind()
            df_list.append(wind_df)
        if self.config["wave_status"]:
            wave_df = self.parse_wave()
            df_list.append(wave_df)
        if self.config["current_status"]:
            current_df = self.parse_current()
            df_list.append(current_df)
        if self.config["water_status"]:
            water_df = self.parse_water()
            df_list.append(water_df)
        # Concatenate all the dataframes (if the list is not empty) into a single dataframe and only in the overlapping period
        if df_list:
            self.data = pd.concat(df_list, axis=1, join="inner")
        print("Parsing data complete!")

    def parse_wind(self):
        """parse_wind [Function to parse the wind input .txt file selected by the user.
        Uses the config attribute to check the correct varaibles are in the input file and names the dataframe Series correspondingly.]

        Returns:
            [pandas.Dataframe]: [Dataframe of the wind data timeseries]
        """
        # Read wind data file into a dataframe
        # wind_file = filedialog.askopenfilename(title="Select the wind data file.")
        wind_file = os.getcwd() + "//wind_data.txt"
        wind_df = pd.read_csv(wind_file, sep="\t", header=None)
        # Check if the number of columns is correct.
        if self.config["10m"]:
            if len(wind_df.columns) != 10:
                sys.exit(
                    "Incorrect number of fields in the wind data file for 10m wind speed = TRUE. Check wind data file or config file and try again."
                )
            wind_df.rename(
                {
                    2: "WS",
                    3: "WnD",
                    4: "T",
                    5: "Roh",
                    6: "WS_10",
                    7: "WnD_10",
                    8: "T_10",
                    9: "Roh_10",
                },
                inplace=True,
                axis="columns",
            )
        else:
            if len(wind_df.columns) != 6:
                sys.exit(
                    "Incorrect number of fields in the wind data file for 10m wind speed = FALSE. Check wind data file or config file and try again."
                )
            wind_df.rename(
                {2: "WS", 3: "WnD", 4: "T", 5: "Roh"}, inplace=True, axis="columns"
            )

        wind_df = make_time_index(wind_df)
        if True in wind_df.index.duplicated():
            sys.exit(
                "Duplicate timestamps in the wind data file. Please check and try again."
            )
        return wind_df

    def parse_wave(self):
        """parse_wave [Function to parse the wave input .txt file selected by the user.
        Uses the config attribute to check the correct varaibles are in the input file and names the dataframe Series correspondingly.]

        Returns:
            [pandas.Dataframe]: [Dataframe of the wave data timeseries]
        """
        # Read wave data file into a dataframe
        # wave_file = filedialog.askopenfilename(title="Select the wave data file.")
        wave_file = os.getcwd() + "//wave_data.txt"
        wave_df = pd.read_csv(wave_file, sep="\t", header=None)
        # Check if there should be spectral wave components (swell and windsea)
        if self.config["wave_spectral"]:
            # Check if the user has input peak enhancement factor.
            if self.config["peak_enhancement"]:
                # Checks the correct number of columns in the wave .txt file
                if len(wave_df.columns) != 17:
                    sys.exit(
                        "Incorrect number of fields in the wave data file for spectral components = TRUE and Peak Enhancement Factor = TRUE. Check wave data file or config file and try again."
                    )
                wave_df.rename(
                    {
                        2: "Hs",
                        3: "WvD",
                        4: "Tp",
                        5: "Tz",
                        6: "G",
                        7: "Hs_W",
                        8: "WvD_W",
                        9: "Tp_W",
                        10: "Tz_W",
                        11: "G_W",
                        12: "Hs_S",
                        13: "WvD_S",
                        14: "Tp_S",
                        15: "Tz_S",
                        16: "G_S",
                    },
                    inplace=True,
                    axis="columns",
                )
            # If no peak enhancement factor is input in the .txt file
            else:
                if len(wave_df.columns) != 14:
                    sys.exit(
                        "Incorrect number of fields in the wave data file for spectral components = TRUE and Peak Enhancement Factor = FALSE. Check wave data file or config file and try again."
                    )
                wave_df.rename(
                    {
                        2: "Hs",
                        3: "WvD",
                        4: "Tp",
                        5: "Tz",
                        6: "Hs_W",
                        7: "WvD_W",
                        8: "Tp_W",
                        9: "Tz_W",
                        10: "Hs_S",
                        11: "WvD_S",
                        12: "Tp_S",
                        13: "Tz_S",
                    },
                    inplace=True,
                    axis="columns",
                )
                if self.config["derive_peak_enhancement"]:
                    wave_df = self.get_gamma(wave_df) #populate wave_df with values for gamma

        # If there are no spectral components.
        else:
            # Check if the user has input peak enhancement factor.
            if self.config["peak_enhancement"]:
                if len(wave_df.columns) != 7:
                    sys.exit(
                        "Incorrect number of fields in the wave data file for spectral components = FALSE and Peak Enhancement Factor = TRUE. Check wave data file or config file and try again."
                    )
                wave_df.rename(
                    {2: "Hs", 3: "WvD", 4: "Tp", 5: "Tz", 6: "G"},
                    inplace=True,
                    axis="columns",
                )
            # If no peak enhancement factor is input in the .txt file
            else:
                if len(wave_df.columns) != 6:
                    sys.exit(
                        "Incorrect number of fields in the wave data file for spectral components = FALSE and Peak Enhancement Factor = FALSE. Check wave data file or config file and try again."
                    )
                wave_df.rename(
                    {2: "Hs", 3: "WvD", 4: "Tp", 5: "Tz"}, inplace=True, axis="columns"
                )
                if self.config["derive_peak_enhancement"]:
                    wave_df = self.get_gamma(wave_df)  #populate wave_df with values for gamma

        wave_df = make_time_index(wave_df)
        if True in wave_df.index.duplicated():
            sys.exit(
                "Duplicate timestamps in the wave data file. Please check and try again."
            )
        return wave_df

    def parse_current(self):
        """parse_current [Function to parse the current input .txt file selected by the user.
        Uses the config attribute to check the correct varaibles are in the input file and names the dataframe Series correspondingly.]

        Returns:
            [pandas.Dataframe]: [Dataframe of the current data timeseries]
        """
        # Read wave data file into a dataframe
        # current_file = filedialog.askopenfilename(title="Select the current data file.")
        current_file = os.getcwd() + "//current_data.txt"

        current_df = pd.read_csv(current_file, sep="\t", header=None)
        # Check if there are tidal and residual current components
        if self.config["current_components"]:
            if len(current_df.columns) != 11:
                sys.exit(
                    "Incorrect number of fields in the current data file for current components = TRUE. Check current data file or config file and try again."
                )
            current_df.rename(
                {
                    2: "SV",
                    3: "DaV",
                    4: "CD",
                    5: "SV_Tid",
                    6: "Dav_Tid",
                    7: "CD_Tid",
                    8: "SV_Res",
                    9: "DaV_Res",
                    10: "CD_Res",
                },
                inplace=True,
                axis="columns",
            )
        else:
            if len(current_df.columns) != 5:
                sys.exit(
                    "Incorrect number of fields in the current data file for current components = FALSE. Check current data file or config file and try again."
                )
            current_df.rename(
                {2: "SV", 3: "DaV", 4: "CD"}, inplace=True, axis="columns"
            )
        current_df = make_time_index(current_df)
        if True in current_df.index.duplicated():
            sys.exit(
                "Duplicate timestamps in the current data file. Please check and try again."
            )
        return current_df

    def parse_water(self):
        """parse_wiater [Function to parse the water input .txt file selected by the user.
        Uses the config attribute to check the correct varaibles are in the input file and names the dataframe Series correspondingly.]

        Returns:
            [pandas.Dataframe]: [Dataframe of the water data timeseries]
        """
        # Read water data file into a dataframe
        # water_file = filedialog.askopenfilename(title="Select the seawater data file.")
        water_file = os.getcwd() + "//water_data.txt"

        water_df = pd.read_csv(water_file, sep="\t", header=None)
        # Check if the water file has the correct number of columns.
        if len(water_df.columns) != 5:
            sys.exit(
                "Incorrect number of field in the water file. Check water data file of config file and try again."
            )
        water_df.rename({2: "Salt", 3: "SST", 4: "Roh_W"}, inplace=True, axis="columns")
        water_df = make_time_index(water_df)
        if True in water_df.index.duplicated():
            sys.exit(
                "Duplicate timestamps in the water data file. Please check and try again."
            )
        return water_df
    
    def get_gamma(self, wave_df):

        wave_df["G"] = (wave_df["Tp"]/np.sqrt(wave_df["Hs"])).map(gamma_DNVGL)

        if self.config["wave_spectral"]:
            wave_df["G_W"] = (wave_df["Tp_W"]/np.sqrt(wave_df["Hs_W"])).map(gamma_DNVGL)
            wave_df["G_S"] = 10

        return wave_df

    def sectorise(self):
        """sectorise [Creates new columns into self.data for all the relevant variables 
        which need to be divided into bins or sectors. 
        Column headers are the same as the original plus "_bins" or "_sectors"]
        """     
        right = False
        if self.config["bin_type"] == "right": right = True

        if self.config["wind_status"]:
            self.data["WS_bins"] = self.get_bins("WS", self.config["wind_bin_size"], right)
            self.data["WnD_sectors"] = self.get_sectors("WnD", self.config["wind_sectors"], right)
            if self.config["10m"]:
                self.data["WS_10_bins"] = self.get_bins("WS_10", self.config["wind_bin_size"], right)
                self.data["WnD_10_sectors"] = self.get_sectors("WnD_10", self.config["wind_sectors"], right)

        if self.config["wave_status"]:
            self.data["Hs_bins"] = self.get_bins("Hs", self.config["wave_height_bin_size"], right)
            self.data["Tp_bins"] = self.get_bins("Tp", self.config["wave_period_bin_size"], right)
            self.data["Tz_bins"] = self.get_bins("Tz", self.config["wave_period_bin_size"], right)
            self.data["WvD_sectors"] = self.get_sectors("WvD", self.config["wave_sectors"], right)

            if self.config["wave_spectral"]:
                self.data["Hs_W_bins"] = self.get_bins("Hs_W", self.config["wave_height_bin_size"], right)
                self.data["Tp_W_bins"] = self.get_bins("Tp_W", self.config["wave_period_bin_size"], right)
                self.data["Tz_W_bins"] = self.get_bins("Tz_W", self.config["wave_period_bin_size"], right)
                self.data["WvD_W_sectors"] = self.get_sectors("WvD_W", self.config["wave_sectors"], right)
                self.data["Hs_S_bins"] = self.get_bins("Hs_S", self.config["wave_height_bin_size"], right)
                self.data["Tp_S_bins"] = self.get_bins("Tp_S", self.config["wave_period_bin_size"], right)
                self.data["Tz_S_bins"] = self.get_bins("Tz_S", self.config["wave_period_bin_size"], right)
                self.data["WvD_S_sectors"] = self.get_sectors("WvD_S", self.config["wave_sectors"], right)

            if self.config["current_status"]:
                self.data["SV_bins"] = self.get_bins("SV", self.config["current_bin_size"], right)
                self.data["DaV_bins"] = self.get_bins("DaV", self.config["current_bin_size"], right)
                self.data["CD_sectors"] = self.get_sectors("CD", self.config["current_sectors"], right)
                if self.config["current_components"]:
                    self.data["SV_Tid_bins"] = self.get_bins("SV_Tid", self.config["current_bin_size"], right)
                    self.data["DaV_Tid_bins"] = self.get_bins("DaV_Tid", self.config["current_bin_size"], right)
                    self.data["CD_Tid_sectors"] = self.get_sectors("CD_Tid", self.config["current_sectors"], right)
                    self.data["SV_Res_bins"] = self.get_bins("SV_Res", self.config["current_bin_size"], right)
                    self.data["DaV_Res_bins"] = self.get_bins("DaV_Res", self.config["current_bin_size"], right)
                    self.data["CD_Res_sectors"] = self.get_sectors("CD_Res", self.config["current_sectors"], right)

    def get_bins(self, header, bin_size, right):
        """get_bins [Function to get bin values for a specific column under self.data and populate self.bins]

        Args:
            header ([string]): [header of the column in self.data to get bins from]
            bin_size ([float]): [size of the bins for this variable, as specified in self.config]
            right ([bool]): [indicates if right boundary is closed. If False, left boudnary is closed] 
        
        Returns:
            [list]: [list to append to self.data containing binned values]
        """
        bines = np.arange(0, self.data[str(header)].max(), bin_size)
        bin_list = np.digitize(self.data[str(header)], bins=bines, right=right)*bin_size-bin_size/2
        self.bins[header] = bines + bin_size/2

        return bin_list
    
    def get_sectors(self, header, N_Sectors, right):
        """get_sectors [Function to get sector values for a specific column under self.data]

        Args:
            header ([string]): [header of the column in self.data to get sectors from]
            N_sectors ([int]): [number of sectors for this variable, as specified in self.config]
            right ([bool]): [indicates if right boundary is closed. If False, left boudnary is closed] 
        
        Returns:
            [list]: [list to append to self.data containing sectorised values]
        """
        if right:
            sector_list = np.where(
                self.data[header] > (360 - ((360/N_Sectors)/2)),1,
                self.data[header].apply(lambda x: np.ceil(((x/(360/N_Sectors))+0.5))).astype('Int64'))

        else:
            sector_list = np.where(
               self.data[header] >= (360 - ((360/N_Sectors)/2)),1,
               (((self.data[header]+(360/N_Sectors)/2)/(360/N_Sectors))+1).apply(np.floor).astype('Int64'))


        return sector_list

def make_time_index(df):
    """make_time_index Creates a DateTime index for the dataframes read from the user input .txt files in the YYYY-MM-DD HH:MM format. Deletes the YYMMDD and HHMM string columns.

    Args:
        df (pandas.Dataframe): [Timeseries DataFrame input by user. Can be wind, wave, current or seawater dataframe.]

    Returns:
        [pandas.DataFrame]: [Returns the input dataframe with the DateTime index.]
    """
    df.iloc[:, 0] = pd.to_datetime(df.iloc[:, 0], format="%Y%m%d")
    df.iloc[:, 1] = pd.to_timedelta(df.iloc[:, 1] / 100, unit="hours")
    df.index = df.iloc[:, 0] + df.iloc[:, 1]
    df.drop(columns=[0, 1], inplace=True)
    return df

def gamma_DNVGL(x):
    """gamma_DNVGL returns the gamma value (peak enhancement factor) according to the methodology proposed by DNVGL in RP-C205.

    Args:
        x (scalar): [Value to determine peak enhancement factor. The coefficient of Tp over the square root of Hs]

    Returns:
        [Scalar]: [Returns the estimate of the peak enhancement factor as a scalar]
    """
    if np.isnan(x):
        sys.exit(   
                "Erroneous value found in calculation of peak enhancement factor. Possibly a 0 or negative value in Hs data. Please check and try again."
            )
    if x <= 3.6: return 5
    elif x >= 5: return 1
    else: return np.exp((5.75-1.15*x))


def main():
    # root = tk.Tk()
    # root.iconbitmap("OW_logo.ico")
    # # Asks the user to select the config file and stores the full path.
    # filepath = filedialog.askopenfilename(
    #     title="Select the metocean configuration file."
    # )
    filepath = os.getcwd() + "//Metocean-BoD_Config Sheet_v0.xlsx"
    metocean_data = MetoceanData(filepath)
    print(metocean_data.data.head())
    #print(list(metocean_data.data))
    # Method for taking mean or median within bin to be implemented 
    if (metocean_data.config["wind_status"] & metocean_data.config["wave_status"]):
        NSS_tables = NSS.NSS(metocean_data)

    print("end")

if __name__ == "__main__":
    main()
