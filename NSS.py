"""
Module for the NSS class
By Juan Teruel
Metocean & Energy Assessment Department
21/12/2020
"""

import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.styles import NamedStyle, Border, Color, Font, Alignment, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import utils

class NSS():    
    """ A class to calculate and print NSS tables from an instance of the MetoceanData object."""

    # Initialise the NSS object using an instance of the MetoceanData object
    def __init__(self, metocean_data):
        # Initialise attributes of NSS object by taking informtation from the metocean_data instance
        print("Calculating NSS tables...")
        self.set_up(metocean_data)
        # Select the relevant data from the metocean_data.data attribute
        self.parse_data(metocean_data)
        # Use the selected data to calculate the NSS tables
        self.get_NSS_tables()
        # Print the NSS tables to excel files
        self.produce_NSS_Excel()

    def set_up(self, metocean_data):
        """set_up: [Initialises the attributes of NSS from information contained in the MetoceanData object]

        Args:
            metoecan_data: [An instance of the MetoceanData object]
        """
        # doesnt check for wind and wave status bc shouldn't be called if they're FALSE
        # TODO: convert to Dictionary
        self.PID = metocean_data.config["project"]
        self.NSectors_wind = metocean_data.config["wind_sectors"]
        self.WS_bins_list = metocean_data.bins["WS"] 
        self.WS_bin_size = metocean_data.config["wind_bin_size"]
        self.WS_HH = metocean_data.config["hub_height"]
        self.NSectors_wave = metocean_data.config["wave_sectors"]
        self.peak_enhancement = metocean_data.config["peak_enhancement"]
        self.derive_peak_enhancement = metocean_data.config["derive_peak_enhancement"]
        self.method = metocean_data.config["method"]
        self.wave_spectral = metocean_data.config["wave_spectral"]
        self.Total_Count = metocean_data.data.shape[0]
        self.closed_boundary = metocean_data.config["bin_type"]

        # Create empty data attribute where to store wind and wave data conviniently. 
        # Create empty tables attribute of the right size to populate afterwards
        self.Total_data = []
        self.Total_tables = np.empty((self.NSectors_wind + 1,self.NSectors_wave + 1,self.WS_bins_list.size,4))
        # If Wind and Swell wave data is included in the MetoceanData object, create data and tables attributes for them too 
        if self.wave_spectral: 
            self.Wind_data, self.Swell_data = [],[]
            self.Wind_tables = np.empty((self.NSectors_wind + 1,self.NSectors_wave + 1,self.WS_bins_list.size,4))
            self.Swell_tables = np.empty((self.NSectors_wind + 1,self.NSectors_wave + 1,self.WS_bins_list.size,4)) # Swell sea not impacted by Wind Direction

    def parse_data(self, metocean_data):
        """parse_data: [Populates the data attributes with wind and wave data taken from the MetoceanData object]

        Args:
            metoecan_data: [An instance of the MetoceanData object]
        """
        # Total_data attribute is always present
        if self.peak_enhancement == False and self.derive_peak_enhancement == False:
            self.Total_data = metocean_data.data.loc[:,("WS_bins","WnD_sectors","WvD_sectors","Hs","Tp")]
            self.Total_data["G"] = np.NAN
        else:
           self.Total_data = metocean_data.data.loc[:,("WS_bins","WnD_sectors","WvD_sectors","Hs","Tp","G")]

        # If Wind and Swell data are to be included, populate their respective attributes 
        # and rename their variables for convinient handling 
        if self.wave_spectral:
            if self.peak_enhancement == False and self.derive_peak_enhancement == False:
                self.Wind_data = metocean_data.data.loc[:,("WS_bins","WnD_sectors","WvD_W_sectors","Hs_W","Tp_W")]
                self.Wind_data["G_W"] = np.NAN
                self.Swell_data = metocean_data.data.loc[:,("WS_bins","WnD_sectors","WvD_S_sectors","Hs_S","Tp_S")]
                self.Swell_data["G_S"] = np.NAN
            else:
                self.Wind_data = metocean_data.data.loc[:,("WS_bins","WnD_sectors","WvD_W_sectors","Hs_W","Tp_W","G_W")]
                self.Swell_data = metocean_data.data.loc[:,("WS_bins","WnD_sectors","WvD_S_sectors","Hs_S","Tp_S","G_S")]
            self.Wind_data.rename(
                columns={"WvD_W_sectors": "WvD_sectors","Hs_W": "Hs","Tp_W":"Tp","G_W":"G"}, inplace=True)
            self.Swell_data.rename(
                columns={"WvD_S_sectors": "WvD_sectors","Hs_S": "Hs","Tp_S":"Tp","G_S":"G"}, inplace=True)

    def get_NSS_tables(self):
        """get_NSS_tables: [Populates the tables attributes]

           Tables are uniform in size and containt 4 dimensions, for:
            1. Wind Direction Sectors
            2. Wave Direction Sectors
            3. Wind Speed bins. Empty wind speed bins are populated with NaNs
            4. Hs, Tp, Peak enhancement factor and Probability of ocurrence

        """ 
        print("Calculating NSS tables...")      
        # Calculate tables for NSS Total Sea and populate NSS.Total_tables attribute
        for WnSector in range(0,self.NSectors_wind + 1):
            for WvSector in range(0,self.NSectors_wave + 1):
                if WnSector == 0: # OMNIDIRECTIONAL
                    if WvSector == 0: # OMNIDIRECTIONAL
                        self.Total_tables[WnSector][WvSector] = self.calc_table(self.Total_data)
                    else:
                        df_temp =  self.Total_data[self.Total_data.WvD_sectors == WvSector]
                        self.Total_tables[WnSector][WvSector] = self.calc_table(df_temp)
                else:
                    if WvSector == 0:
                        df_temp =  self.Total_data[self.Total_data.WnD_sectors == WnSector]
                        self.Total_tables[WnSector][WvSector] = self.calc_table(df_temp)
                    else:
                        df_temp = self.Total_data[
                            (self.Total_data.WvD_sectors == WvSector) & (self.Total_data.WnD_sectors == WnSector)]
                        self.Total_tables[WnSector][WvSector] = self.calc_table(df_temp)        

        # Calculate tables for NSS Wind and Swell Sea
        if self.wave_spectral:
            print("Boiling virtual kettle for virtual tea...")    
            for WnSector in range(0,self.NSectors_wind + 1):
                for WvSector in range(0,self.NSectors_wave + 1):
                    if WnSector == 0: 
                        if WvSector == 0: 
                            self.Swell_tables[WnSector][WvSector] = self.calc_table(self.Swell_data)
                            self.Wind_tables[WnSector][WvSector] = self.calc_table(self.Wind_data)
                        else:
                            df_temp =  self.Swell_data[self.Swell_data.WvD_sectors == WvSector]
                            self.Swell_tables[WnSector][WvSector] = self.calc_table(df_temp)
                            df_temp =  self.Wind_data[self.Wind_data.WvD_sectors == WvSector]
                            self.Wind_tables[WnSector][WvSector] = self.calc_table(df_temp)
                    else: # SWELL COMPONENT SHOULDNT BE AFFECTED BY WIND, BUT INCLUDED ATM
                        if WvSector == 0: # Wind tables contain values for filtered wind but omnidirectional waves
                            df_temp = self.Swell_data[self.Swell_data.WnD_sectors == WnSector]
                            self.Swell_tables[WnSector][WvSector] = self.calc_table(df_temp)
                            df_temp = self.Wind_data[self.Wind_data.WnD_sectors == WnSector]
                            self.Wind_tables[WnSector][WvSector] = self.calc_table(df_temp)
                        else:
                            df_temp = self.Swell_data[
                                (self.Swell_data.WvD_sectors == WvSector) & (self.Swell_data.WnD_sectors == WnSector)]
                            self.Swell_tables[WnSector][WvSector] = self.calc_table(df_temp)
                            df_temp = self.Wind_data[
                                (self.Wind_data.WvD_sectors == WvSector) & (self.Wind_data.WnD_sectors == WnSector)]
                            self.Wind_tables[WnSector][WvSector] = self.calc_table(df_temp)

        print("All NSS Tables calculated!")
        print("Preparing Excel report...")

    def calc_table(self, NSS_data):
        """ calc_table: [creates a single NSS table for a specific combination of wind and wave direction sector.
                    Works the same for Total, Wind or Swell waves.]

            Args: 
                NSS_data ([pandas Dataframe]): a dataframe containing wind and wave data for this particular table,
                    already filtered by wind and wave sector

            Returns:
                tab ([numpy array]): numpy array containing the NSS table
        """
        # Create an empty table of the right size
        tab = np.zeros((self.WS_bins_list.size,4))

        # Iterate and populate with calculated values or NaNs if there are no registries for a particular wind speed bin
        for i in range(self.WS_bins_list.size):
            bin_centre = self.WS_bins_list[i]
            df_temp = NSS_data[NSS_data.WS_bins == bin_centre]
            if df_temp.shape[0] == 0:
                tab[i] = [np.NAN, np.NAN, np.NAN, np.NAN]
            else:
                # probability of ocurrence of this particular wind speed bin in this particular wind and wave direction sector combination
                # over the total number of events in the timeseries
                tab[i][3] = df_temp["Hs"].size/self.Total_Count

                # Mean or median depending on user selection
                if self.method == "mean":
                    tab[i][0] = df_temp["Hs"].mean()
                    tab[i][1] = df_temp["Tp"].mean()
                    tab[i][2] = df_temp["G"].mean()
                        
                else:
                    tab[i][0] = df_temp["Hs"].median()
                    tab[i][1] = df_temp["Tp"].median()
                    tab[i][2] = df_temp["G"].median()

        return tab

    def produce_NSS_Excel(self):
        """ produce_NSS_Excel: [routine to produce ant Excel .xlsx file which contains the NSS tables fully formatted]"""

        wb = Workbook()
        create_styles(wb)

        ws_Total = wb.active
        ws_Total.title = "NSS Total sea"
        #ws_Total.sheet_properties.tabColor = "072B31"
        ws_Total.sheet_view.showGridLines = False

        # create table with the WS bins
        self.WS_bin_table = np.stack((
            self.WS_bins_list - self.WS_bin_size/2,
            self.WS_bins_list,
            self.WS_bins_list + self.WS_bin_size/2), axis=1
            ) 
        # Dictionary containing specific information for the tables.
        self.NSS_table_info = {
            "WS_info": "Hourly Mean WS at {}mMSL [m/s]".format(self.WS_HH),
            "NSS Total sea": 1,
            "NSS Wind sea": 2,
            "NSS Swell sea": 3
        }
        # Determine appropiate headers for the WS bins
        if self.closed_boundary == "left":
            self.WS_bin_headers = ["Lower (>=)","Middle","Upper (<)"]
        elif self.closed_boundary == "right":
            self.WS_bin_headers = ["Lower (>)","Middle","Upper (<=)"]
        self.NSS_table_headers = ["Hs [m]","Tp [s]","γ [-]","Prob [%]"]

        # Call print routine for the Total tables
        self.print_NSS_tables(ws_Total, self.Total_tables, 2, 2)

        # If required, call print routine also for the Wind and Swell Sea tables
        if self.wave_spectral:
            print("Putting on favourite tune for motivation...")
            ws_Wind = wb.create_sheet("NSS Wind sea", 1)
            #ws_Wind.sheet_properties.tabColor = "D9D9D6"  
            ws_Wind.sheet_view.showGridLines = False
            self.print_NSS_tables(ws_Wind, self.Wind_tables, 2, 2)

            ws_Swell = wb.create_sheet("NSS Swell sea", 2)
            #ws_Swell.sheet_properties.tabColor = "D9D9D6"
            ws_Swell.sheet_view.showGridLines = False
            self.print_NSS_tables(ws_Swell, self.Swell_tables, 2, 2)
        
        wb.save("{}_Metocean_NSS_Tables.xlsx".format(self.PID)) 
        print("Excel report complete!")   

    def print_NSS_tables(self, ws, data, startRow, startCol):
        """ print_NSS_tables: [writes the NSS tables to the target worksheet]

            Args: 
                ws ([openpyxl worksheet object]): the worksheet to write results to
                data ([numpy array]): table containing the data to write. Expects full table (i.e. Total, Wind or Swell sea)
                startRow ([integer]): row where to start printing the results
                startCol ([integer]): column where to start printing the results

        """

        index_titles = [self.NSS_table_info["WS_info"], "Wind Sector", "Wave Sector"]
        # Uses title of the worksheet to determine first digit of the table numbers
        table_number = self.NSS_table_info[ws.title] 

        # OMNI-OMNI
        print_table(ws, self.WS_bin_table, index_titles, self.WS_bin_headers, startRow, startCol, "NSS_index")
        col = startCol + 3
        table_titles = ["Table {}.0.0".format(table_number), "OMNI", "OMNI"]
        print_table(ws, data[0][0], table_titles, self.NSS_table_headers, startRow, col, "conditional")    
        startRow += len(index_titles) + len(self.WS_bins_list) + 4

        # Sect-OMNI
        print_table(ws, self.WS_bin_table, index_titles, self.WS_bin_headers, startRow, startCol, "NSS_index")
        col = startCol + 3
        for WnSector in range(1, data.shape[0]):
            table_titles = ["Table {}.{}.0".format(table_number, WnSector), WnSector, "OMNI"]
            print_table(ws, data[WnSector][0], table_titles, self.NSS_table_headers, startRow, col, "conditional")
            col += 4
        startRow += len(index_titles) + len(self.WS_bins_list) + 4

        # OMNI-Sect
        print_table(ws, self.WS_bin_table, index_titles, self.WS_bin_headers, startRow, startCol, "NSS_index")
        col = startCol + 3
        for WvSector in range(1, data.shape[1]):
            table_titles = ["Table {}.0.{}".format(table_number, WvSector), "OMNI", WvSector]
            print_table(ws, data[0][WvSector], table_titles, self.NSS_table_headers, startRow, col, "conditional")
            col += 4
        
        # Sect-Sect
        for WnSector in range(1, data.shape[0]):
            startRow += len(index_titles) + len(self.WS_bins_list) + 4
            print_table(ws, self.WS_bin_table, index_titles, self.WS_bin_headers, startRow, startCol, "NSS_index")
            col = startCol + 3
            for WvSector in range(1, data.shape[1]):
                titles = [
                    "Table {}.{}.{}".format(table_number,WnSector, WvSector), WnSector, WvSector]           
                print_table(ws, data[WnSector][WvSector], titles, self.NSS_table_headers, startRow, col, "conditional")
                col += 4

def create_styles(wb):
        """ crate_styles: [create styles to the target workbook object]

            Args: 
                wb ([openpyxl workbook object]): workbook where to create the styles

        """
    NSS_header = NamedStyle(name="NSS_header")
    NSS_header.fill = PatternFill(fill_type="solid", fgColor="072B31")
    NSS_header.font = Font(bold=True, color="00FFFFFF")
    NSS_header.alignment = Alignment(horizontal="center")
    wb.add_named_style(NSS_header)

    NSS_index = NamedStyle(name="NSS_index")
    NSS_index.font = Font(bold=True)
    NSS_index.fill = PatternFill(fill_type="solid", fgColor="D9D9D6")
    NSS_index.alignment = Alignment(horizontal="center")
    wb.add_named_style(NSS_index)

    Wind_header = NamedStyle(name="Wind_header")
    Wind_header.font = Font(bold=True)
    Wind_header.fill = PatternFill(fill_type="solid", fgColor="FFE900")
    Wind_header.alignment = Alignment(horizontal="center")
    wb.add_named_style(Wind_header)

    Wave_header = NamedStyle(name="Wave_header")
    Wave_header.font = Font(bold=True)
    Wave_header.fill = PatternFill(fill_type="solid", fgColor="01C1D5")
    Wave_header.alignment = Alignment(horizontal="center")
    wb.add_named_style(Wave_header)

def print_table(ws, data, titles, headers, startRow, startCol, style):
    """ print_table: [writes an individual NSS table to the target worksheet]

        Args: 
            ws ([openpyxl worksheet object]): the worksheet to write results to
            data ([numpy array]): table containing the data to write.
            headers ([list]): list of headers to write above the table. Generally, information to identify the table.
            startRow ([integer]): row where to start printing the results
            startCol ([integer]): column where to start printing the results
            style ([string]): stlye to apply to the table. If conditional, data will be apply conditional formatting per column.
                    I           If a style, it must already exist in the workbook.

    """
    rows = data.shape[0]
    cols = data.shape[1]
    prob_sum = 0

    for t in range(len(titles)):
        ws.merge_cells(start_row = startRow + t,
                    start_column = startCol,
                    end_row = startRow + t,
                    end_column = startCol + cols - 1)
        ws.cell(startRow + t, startCol).value = titles[t]
        if t == 0:
            ws.cell(row = startRow + t, column = startCol).style = "NSS_header"
        elif titles[t] == "Wind Sector":
            ws.cell(row = startRow + t, column = startCol).style = "Wind_header"
        elif titles[t] == "Wave Sector":
            ws.cell(row = startRow + t, column = startCol).style = "Wave_header"
        else:
            ws.cell(row = startRow + t, column = startCol).style = "NSS_index"
        outside_borders(ws, startRow + t, startCol, startRow + t, startCol + cols - 1)
        
    
    startRow += len(titles)

    for h in range(len(headers)):
        ws.cell(row = startRow, column = startCol + h).value = headers[h]
        ws.cell(row = startRow, column = startCol + h).style = "NSS_header"

    outside_borders(ws, startRow, startCol, startRow, startCol + len(headers) - 1)
    startRow += 1

    for r in range(rows):
        row = startRow + r
        for c in range(cols):
            col = startCol + c
            if data.dtype == np.float and np.isnan(data[r][c]):
                ws.cell(row = row, column = col).value = "NaN"
                ws.cell(row = row, column = col).alignment = Alignment(horizontal="center")
            else:
                if data[r][c] != "":
                    ws.cell(row = row, column = col).value = np.float(data[r][c])
                    ws.cell(row = row, column = col).alignment = Alignment(horizontal="center")
                    if cols == 3:
                        ws.cell(row = row, column = col).style = style
                    else:
                        if c == 3:
                            ws.cell(row = row, column = col).number_format = "0.00%"
                            prob_sum += data[r][c] 
                        else:
                            ws.cell(row = row, column = col).number_format = "0.00"   
    
    if style == "conditional": 
        for c in range(cols): 
            col = startCol + c
            col_letter = utils.cell.get_column_letter(col)
            col_range = "{}{}:{}{}".format(col_letter,startRow,col_letter,startRow+rows-1)
            ws.conditional_formatting.add(col_range,
                        ColorScaleRule(start_type="min", start_color="63BE7B",
                                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                        end_type="max", end_color="F8696B")
                        )

    endRow = startRow + rows - 1
    endCol = startCol + cols - 1 
    outside_borders(ws, startRow, startCol, endRow, endCol)

    # Adds footer section. For index, simply add "SUM", for data, add the sum of the table probabilities.
    if cols == 3:
        ws.merge_cells(start_row = endRow + 1,
                        start_column = startCol,
                        end_row = endRow + 1,
                        end_column = endCol)
        ws.cell(row=endRow + 1, column=startCol).value = "SUM"                
        ws.cell(row=endRow + 1, column=startCol).style = "NSS_index"
        ws.cell(row=endRow + 1, column=startCol).font = Font(bold=True)
    elif cols == 4:
        ws.cell(row=endRow + 1, column=startCol+3).value = prob_sum
        ws.cell(row=endRow + 1, column=startCol+3).alignment = Alignment(horizontal="center")
        ws.cell(row=endRow + 1, column=startCol+3).number_format = "0.00%"
    outside_borders(ws, endRow + 1, startCol, endRow + 1, endCol)

def outside_borders(ws, startRow, startCol, endRow, endCol, style="thin"): 
    """ outside_borders: [draws outside borders for a range in an excel worksheet]

    Args: 
        ws ([openpyxl worksheet object]): the worksheet to edit
        startRow ([integer]): row of the top left cell of the range
        startCol ([integer]): column of the top left cell of the range
        endRow ([integer]): row of the bottom right cell of the range
        endCol ([integer]): column of the bottom right cell of the range
        style ([string]): border style. Must match those allowed by Excel.

    """
    if startCol == endCol:
        if startRow == endRow:
            ws.cell(row=startRow, column = startCol).border = Border(
                top=Side(style=style),
                bottom=Side(style=style),
                right=Side(style=style),
                left=Side(style=style)
            ) 
        else:
            for row in range(startRow, endRow + 1):
                if row == startRow:
                        ws.cell(row=row, column = startCol).border = Border(
                            top=Side(style=style),
                            right=Side(style=style),
                            left=Side(style=style)
                        )
                elif row == endRow:
                    ws.cell(row=row, column = startCol).border = Border(
                        bottom=Side(style=style),
                        right=Side(style=style),
                        left=Side(style=style)
                    )
                else:
                    ws.cell(row=row, column = startCol).border = Border(
                        right=Side(style=style),
                        left=Side(style=style)
                    )
    else:
        if startRow == endRow:
            for col in range(startCol, endCol + 1):
                if col == startCol:
                        ws.cell(row=startRow, column = col).border = Border(
                            top=Side(style=style),
                            bottom=Side(style=style),
                            left=Side(style=style)
                        )
                elif col == endCol:
                    ws.cell(row=startRow, column = col).border = Border(
                        top=Side(style=style),
                        bottom=Side(style=style),
                        right=Side(style=style),
                    )
                else:
                    ws.cell(row=startRow, column = col).border = Border(
                        top=Side(style=style),
                        bottom=Side(style=style),
                    )
        else:
            for col in range(startCol, endCol + 1):
                for row in range(startRow, endRow + 1):
                    if col == startCol:
                        if row == startRow:
                            ws.cell(row=row, column = col).border = Border(
                                top=Side(style=style),
                                left=Side(style=style)
                            )
                        elif row == endRow:
                            ws.cell(row=row, column = col).border = Border(
                                bottom=Side(style=style),
                                left=Side(style=style)
                            )
                        else:
                            ws.cell(row=row, column = col).border = Border(
                                left=Side(style=style)
                            )
                    elif col == endCol:
                        if row == startRow:
                            ws.cell(row=row, column = col).border = Border(
                                top=Side(style=style),
                                right=Side(style=style)
                            )
                        elif row == endRow:
                            ws.cell(row=row, column = col).border = Border(
                                bottom=Side(style=style),
                                right=Side(style=style)
                            )
                        else:
                            ws.cell(row=row, column = col).border = Border(
                                right=Side(style=style)
                            )
                    else:
                        if row == startRow:
                            ws.cell(row=row, column = col).border = Border(
                                top=Side(style=style),
                            )
                        elif row == endRow:
                            ws.cell(row=row, column = col).border = Border(
                                bottom=Side(style=style),
                            ) 